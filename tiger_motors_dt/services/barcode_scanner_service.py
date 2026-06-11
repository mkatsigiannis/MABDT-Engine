"""
Tiger Motors Digital Twin - Simple Barcode Scanner Service

Simplified implementation that mirrors the original zeroconf_service.py:
- Zeroconf service discovery
- Independent MQTT client
- Simple one-message-at-a-time Excel logging
- All MQTT topics logged to separate sheets

Based on the proven, working original implementation.
"""

import json
import os
import socket
import struct
import threading
import time
from datetime import datetime
from typing import Any

import openpyxl
import paho.mqtt.client as mqtt
from openpyxl import Workbook
from PySide6.QtCore import QObject, Signal
from zeroconf import ServiceInfo, Zeroconf

from mabdt.utils.logging import get_logger
from tiger_motors_dt.config import load_config

logger = get_logger(__name__)


class BarcodeDataLogger:
    """Simple Excel logger - mirrors original zeroconf_service.py"""

    def __init__(self, base_file_path: str = "diagrams_and_data"):
        """Initialize the data logger."""
        self.base_file_path = base_file_path
        self.ensure_directory_exists()

        # Create timestamped file name
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.log_file_name = os.path.join(
            self.base_file_path, f"TigerMotors_MQTT_Log_{timestamp}.xlsx"
        )

        # Track message counts for statistics
        self.message_counts = {}
        self.start_time = datetime.now()

        # Thread lock for file access to prevent corruption
        self._file_lock = threading.Lock()

        # Track if we need to create initial file
        self._file_initialized = False

    def ensure_directory_exists(self):
        """Create the data directory if it doesn't exist."""
        if not os.path.exists(self.base_file_path):
            os.makedirs(self.base_file_path)

    def create_or_update_excel(self, file_name: str, sheet_name: str, data: list):
        """
        Thread-safe Excel write with proper file handling.

        Args:
            file_name: Path to Excel file
            sheet_name: Name of the sheet to update
            data: List of rows to add to the sheet
        """
        # Use lock to prevent concurrent access
        with self._file_lock:
            workbook = None
            try:
                # Check if the file exists
                if os.path.exists(file_name):
                    try:
                        workbook = openpyxl.load_workbook(file_name)
                    except Exception as load_error:
                        # If file is corrupted, create backup and start fresh
                        logger.error(f"Corrupted file detected, creating backup: {load_error}")
                        backup_name = file_name.replace(
                            ".xlsx", f'_corrupted_{datetime.now().strftime("%H%M%S")}.xlsx'
                        )
                        try:
                            os.rename(file_name, backup_name)
                            logger.info(f"Backup saved as: {backup_name}")
                        except OSError as rename_error:
                            logger.warning(f"Could not rename corrupted file: {rename_error}")
                        workbook = Workbook()
                else:
                    workbook = Workbook()
                    # Remove default sheet if it exists and is empty
                    if "Sheet" in workbook.sheetnames:
                        default_sheet = workbook["Sheet"]
                        if default_sheet.max_row == 1 and default_sheet.cell(1, 1).value is None:
                            workbook.remove(default_sheet)

                # Check if the sheet exists
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                else:
                    sheet = workbook.create_sheet(title=sheet_name)

                # Find the next available row
                if sheet.max_row == 1 and sheet.cell(row=1, column=1).value is None:
                    next_row = 1
                else:
                    next_row = sheet.max_row + 1

                # Write data rows (only first two columns like original)
                for row_data in data:
                    for col_index, cell_value in enumerate(row_data, start=1):
                        if col_index <= 2:  # Only timestamp and message
                            sheet.cell(row=next_row, column=col_index, value=cell_value)
                    next_row += 1

                # Save the workbook with error handling
                try:
                    workbook.save(file_name)
                except PermissionError:
                    logger.error("File is open in Excel - cannot save. Data will be lost.")
                except Exception as save_error:
                    logger.error(f"Error saving Excel file: {save_error}")
                    raise

            except Exception as e:
                # Error logging - don't let it crash the service
                logger.exception(f"Error writing to Excel: {e}")
            finally:
                # Explicitly close the workbook to release file handles
                if workbook is not None:
                    try:
                        workbook.close()
                    except Exception as close_error:
                        logger.warning(f"Error closing workbook: {close_error}")

    def sanitize_sheet_name(self, topic: str) -> str:
        """Sanitize topic name for use as Excel sheet name."""
        import re

        if not topic:
            return "Empty_Topic"

        # Replace invalid characters for Excel sheet names
        invalid_chars = ["/", "\\", "*", "?", ":", "[", "]"]
        safe_name = topic
        for char in invalid_chars:
            safe_name = safe_name.replace(char, "_")

        # Replace any non-printable characters
        safe_name = re.sub(r"[^\w\-_. ]", "_", safe_name)
        safe_name = safe_name.strip("'_ ")
        safe_name = re.sub(r"_+", "_", safe_name)

        # Limit length to Excel's 31 character limit
        if len(safe_name) > 31:
            safe_name = safe_name[:28] + "..."

        if not safe_name:
            safe_name = "Topic"

        return safe_name

    def log_message(self, topic: str, message: str, timestamp: str | None = None):
        """Log MQTT message to Excel - simple, one at a time like the original."""
        if timestamp is None:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")

        # Count messages per topic
        if topic not in self.message_counts:
            self.message_counts[topic] = 0
        self.message_counts[topic] += 1

        # Create data row
        data_row = [timestamp, message]

        # Sanitize topic name for use as sheet name
        sheet_name = self.sanitize_sheet_name(topic)

        # Write to Excel immediately (like original)
        self.create_or_update_excel(self.log_file_name, sheet_name, [data_row])

    def get_statistics(self) -> dict[str, Any]:
        """Get logging statistics."""
        runtime = datetime.now() - self.start_time
        total_messages = sum(self.message_counts.values())

        # Calculate logged vs filtered messages
        scanner_messages = sum(
            count for topic, count in self.message_counts.items() if topic.startswith("scanner/")
        )
        cycle_time_messages = sum(
            count
            for topic, count in self.message_counts.items()
            if topic.startswith("ws_cycle_time/")
        )
        filtered_messages = sum(
            count
            for topic, count in self.message_counts.items()
            if topic.startswith("leds/") or topic.startswith("ws_init/")
        )
        logged_messages = total_messages - filtered_messages

        return {
            "start_time": self.start_time,
            "runtime": runtime,
            "total_messages": total_messages,
            "logged_messages": logged_messages,
            "filtered_messages": filtered_messages,
            "scanner_messages": scanner_messages,
            "cycle_time_messages": cycle_time_messages,
            "general_messages": total_messages - scanner_messages,
            "topics": len(self.message_counts),
            "message_counts": self.message_counts.copy(),
            "log_file": self.log_file_name,
            "health_status": "healthy",  # Simplified - always healthy
        }

    def cleanup(self):
        """Clean up resources and ensure file is properly closed."""
        with self._file_lock:
            logger.info(
                f"Final statistics: {sum(self.message_counts.values())} total messages logged"
            )
            logger.info(f"Log file: {self.log_file_name}")
            # File handles are closed in create_or_update_excel's finally block
            # This method is here for future expansion if needed


class BarcodeServerService(QObject):
    """
    Simple Barcode Scanner Service - mirrors original zeroconf_service.py

    - Zeroconf service discovery
    - Independent MQTT client
    - Logs all MQTT topics
    - Simple Excel logging
    """

    # Signals for GUI integration
    service_started = Signal()
    service_stopped = Signal()
    message_received = Signal(str, str)  # topic, message
    error_occurred = Signal(str)

    def __init__(self, config: dict | None = None, bus=None):
        super().__init__()

        self.bus = bus

        # Load config.json as the authoritative source; passed-in config
        # (next block) overrides any specific values the caller cares about.
        self.config = self._load_config()

        # If a config was passed in, merge it (passed config overrides loaded values)
        if config:
            self._merge_config(config)

        # Get service-specific configuration from barcode_scanner_service section
        service_config = self.config.get("barcode_scanner_service", {})

        # Use service-specific settings, falling back to mqtt section for backwards compatibility
        mqtt_config = self.config.get("mqtt", {})
        self.host = service_config.get("host", mqtt_config.get("host", "localhost"))
        self.port = service_config.get("port", mqtt_config.get("port", 1883))

        # Service-specific settings
        self.enabled = service_config.get("enabled", True)
        self.service_name = service_config.get("service_name", "BarcodeScanner")
        self.service_type = service_config.get("service_type", "_scanner._tcp.local.")
        self.log_directory = service_config.get("log_directory", "diagrams_and_data")
        self.log_all_topics = service_config.get("log_all_topics", True)
        self.auto_start = service_config.get("auto_start", True)

        logger.info(f"Barcode Scanner Service configured: {self.host}:{self.port}")
        logger.info(f"  enabled={self.enabled}, log_directory={self.log_directory}")

        # Initialize components with configured log directory
        self.data_logger = BarcodeDataLogger(base_file_path=self.log_directory)
        self.mqtt_client = mqtt.Client(
            callback_api_version=mqtt.CallbackAPIVersion.VERSION1,
            client_id="barcode_scanner_service",
        )
        self.mqtt_client.on_connect = self._on_connect_mqtt
        self.mqtt_client.on_message = self._on_message_mqtt
        self.mqtt_client.on_disconnect = self._on_disconnect_mqtt

        self.zeroconf = None
        self.service_info = None
        self.is_running = False
        self.mqtt_connected = False

    def _load_config(self) -> dict:
        """Load configuration from `config.json`. Returns `{}` on failure."""
        try:
            return load_config()
        except Exception as e:
            logger.warning(f"Could not load config.json: {e}")
            return {}

    def _merge_config(self, override_config: dict) -> None:
        """
        Merge override config into the loaded config.
        Only updates values that are explicitly provided in override_config.
        """
        for key, value in override_config.items():
            if (
                isinstance(value, dict)
                and key in self.config
                and isinstance(self.config[key], dict)
            ):
                # Deep merge for nested dicts
                self.config[key].update(value)
            else:
                self.config[key] = value

    def set_event_bus(self, bus):
        """Set EventBus reference for late binding."""
        self.bus = bus
        logger.info("Barcode Scanner Service connected to EventBus")

    def _on_connect_mqtt(self, client, userdata, flags, rc):
        """MQTT connection callback."""
        if rc == 0:
            logger.info("Barcode Scanner Service MQTT connected")
            self.mqtt_connected = True
            # Subscribe to ALL topics
            client.subscribe("#")
            logger.info("Subscribed to all MQTT topics (#)")
        else:
            error_msg = f"MQTT connection failed with code {rc}"
            logger.error(f"{error_msg}")
            self.error_occurred.emit(error_msg)

    def _on_disconnect_mqtt(self, client, userdata, rc):
        """MQTT disconnection callback."""
        self.mqtt_connected = False
        if rc != 0:
            logger.warning(f"Unexpected MQTT disconnection (code: {rc})")

    def _on_message_mqtt(self, client, userdata, msg):
        """MQTT message callback - log and emit."""
        try:
            topic = msg.topic

            # Special handling for binary cycle time data
            if topic.startswith("ws_cycle_time/"):
                message = self._parse_cycle_time_binary(msg.payload)
            else:
                # Decode with error handling for invalid UTF-8 bytes
                try:
                    message = msg.payload.decode("utf-8")
                except UnicodeDecodeError:
                    # Handle invalid UTF-8 - use replacement character
                    message = msg.payload.decode("utf-8", errors="replace")
                    logger.warning(f"Invalid UTF-8 in message from {topic}")

            # Try to extract timestamp from message payload (if JSON with timestamp field)
            # Otherwise use received time
            timestamp = self._extract_timestamp(message)

            # Log to console
            logger.info(f"[{timestamp[:19]}] {topic}: {message}")

            # Filter out noisy topics from Excel logging (but still print them)
            # - leds/ and ws_init/ send too many messages and slow down Excel
            if not topic.startswith("leds/") and not topic.startswith("ws_init/"):
                # Sanitize message content for Excel (remove illegal XML characters)
                sanitized_message = self._sanitize_excel_content(message)
                # Log to Excel
                self.data_logger.log_message(topic, sanitized_message, timestamp)

            # Emit signal for GUI (all messages)
            self.message_received.emit(topic, message)

        except Exception as e:
            error_msg = f"Error processing MQTT message: {e}"
            logger.error(f"{error_msg}")
            self.error_occurred.emit(error_msg)

    def _parse_cycle_time_binary(self, payload: bytes) -> str:
        """
        Parse binary cycle time data from ws_cycle_time/* topics.

        The workstation agent packs cycle time as big-endian double:
        struct.pack('>d', cycle_time_precise)

        Args:
            payload: Raw binary payload (8 bytes)

        Returns:
            Formatted string with cycle time value
        """
        try:
            # Unpack big-endian double (8 bytes)
            cycle_time = struct.unpack(">d", payload)[0]
            # Format to 4 decimal places (matches workstation precision)
            return f"{cycle_time:.4f}s"
        except struct.error:
            # If unpacking fails, return error message
            return f"<Invalid binary: {len(payload)} bytes>"
        except Exception as e:
            return f"<Parse error: {e}>"

    def _sanitize_excel_content(self, message: str) -> str:
        """
        Sanitize message content for Excel compatibility.

        Removes illegal XML characters that Excel doesn't support:
        - Control characters (0x00-0x1F except tab, newline, carriage return)
        - Invalid Unicode characters

        Args:
            message: Original message content

        Returns:
            Sanitized message safe for Excel
        """
        import re

        # Excel doesn't allow these control characters in cells
        # Keep: tab (0x09), newline (0x0A), carriage return (0x0D)
        # Remove: all other control characters (0x00-0x1F)
        illegal_chars = (
            r"[\x00-\x08]"  # Control chars before tab
            r"|[\x0B-\x0C]"  # Control chars between newline and CR
            r"|[\x0E-\x1F]"  # Control chars after CR
        )

        # Remove illegal characters
        sanitized = re.sub(illegal_chars, "", message)

        # Also remove the Unicode replacement character if it's causing issues
        sanitized = sanitized.replace("�", "?")

        return sanitized

    def _extract_timestamp(self, message: str) -> str:
        """
        Extract timestamp from message payload if available, otherwise use current time.

        Checks for JSON messages with 'timestamp' field in these formats:
        - ISO format: "2025-11-04T02:10:02.519964+00:00"
        - Standard format: "2025-11-03 20:10:02"

        Args:
            message: The decoded message payload

        Returns:
            Timestamp string in format 'YYYY-MM-DD HH:MM:SS.ffffff'
        """
        try:
            # Try to parse as JSON
            data = json.loads(message)

            # Check if timestamp field exists
            if isinstance(data, dict) and "timestamp" in data:
                ts_str = data["timestamp"]

                # Try to parse ISO format timestamp
                try:
                    # Handle ISO format with timezone: 2025-11-04T02:10:02.519964+00:00
                    if "T" in ts_str:
                        dt = datetime.fromisoformat(ts_str.replace("Z", "+00:00"))
                        return dt.strftime("%Y-%m-%d %H:%M:%S.%f")
                except ValueError:
                    pass

                # If ISO parsing failed, return the timestamp as-is if it looks valid
                if len(ts_str) >= 19:  # At least YYYY-MM-DD HH:MM:SS
                    return ts_str

        except (json.JSONDecodeError, ValueError, KeyError):
            # Not JSON or no timestamp field - use current time
            pass

        # Default: use current time (received time)
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")

    def start_service(self):
        """Start the barcode scanner service."""
        # Check if service is enabled
        if not self.enabled:
            logger.info("Barcode Scanner Service is disabled in configuration")
            return

        try:
            logger.info("Starting Barcode Scanner Service...")

            # Start Zeroconf
            self._start_zeroconf()

            # Start MQTT client
            logger.info(f"Connecting to MQTT broker at {self.host}:{self.port}")
            self.mqtt_client.connect(self.host, self.port, 60)
            self.mqtt_client.loop_start()

            self.is_running = True
            self.service_started.emit()
            logger.info("Barcode Scanner Service started successfully")

        except Exception as e:
            error_msg = f"Failed to start service: {e}"
            logger.error(error_msg)
            self.error_occurred.emit(error_msg)

    def stop_service(self):
        """Stop the barcode scanner service."""
        try:
            logger.info("Stopping Barcode Scanner Service...")

            # Stop MQTT client first to prevent new messages
            if self.mqtt_client:
                self.mqtt_client.loop_stop()
                self.mqtt_client.disconnect()
                self.mqtt_connected = False

            # Give a moment for any pending writes to complete
            time.sleep(0.5)

            # Clean up data logger (ensures file is properly closed)
            if self.data_logger:
                self.data_logger.cleanup()

            # Stop Zeroconf
            if self.zeroconf and self.service_info:
                self.zeroconf.unregister_service(self.service_info)
                self.zeroconf.close()
                self.zeroconf = None
                self.service_info = None

            self.is_running = False
            self.service_stopped.emit()
            logger.info("Barcode Scanner Service stopped")

        except Exception as e:
            error_msg = f"Error stopping service: {e}"
            logger.error(error_msg)
            self.error_occurred.emit(error_msg)

    def _start_zeroconf(self):
        """Start Zeroconf service discovery."""
        try:
            self.zeroconf = Zeroconf()
            self.service_info = ServiceInfo(
                self.service_type,
                f"{self.service_name}.{self.service_type}",
                addresses=[socket.inet_aton(self.host)],
                port=self.port,
                properties={},
                server=f"{self.host}.local",
            )

            self.zeroconf.register_service(self.service_info)
            logger.info(
                f"Zeroconf service '{self.service_name}' registered on {self.host}:{self.port}"
            )

        except Exception as e:
            raise Exception(f"Failed to start Zeroconf: {e}") from e

    def get_service_status(self) -> dict[str, Any]:
        """Get current service status."""
        status = {
            "is_running": self.is_running,
            "mqtt_connected": self.mqtt_connected,
            "enabled": self.enabled,
            "host": self.host,
            "port": self.port,
            "service_name": self.service_name,
            "log_directory": self.log_directory,
            "log_all_topics": self.log_all_topics,
        }

        # Add data logger statistics
        if self.data_logger:
            stats = self.data_logger.get_statistics()
            status.update(stats)

        return status


# For standalone testing
if __name__ == "__main__":
    import sys

    from PySide6.QtWidgets import QApplication

    app = QApplication(sys.argv)

    service = BarcodeServerService()
    service.start_service()

    try:
        sys.exit(app.exec())
    except KeyboardInterrupt:
        service.stop_service()
